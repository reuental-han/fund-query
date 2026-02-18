from flask import Flask, jsonify, request, send_file
from flask_cors import CORS
import json
import os
import re
import io
import urllib.request
import urllib.error
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

app = Flask(__name__)
CORS(app)

FUNDS_FILE = 'funds.json'
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

def load_funds():
    funds_path = os.path.join(BASE_DIR, FUNDS_FILE)
    if os.path.exists(funds_path):
        with open(funds_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
            if data and isinstance(data[0], str):
                return [{'code': code, 'shares': None} for code in data]
            return data
    return []

def save_funds(funds):
    funds_path = os.path.join(BASE_DIR, FUNDS_FILE)
    with open(funds_path, 'w', encoding='utf-8') as f:
        json.dump(funds, f, ensure_ascii=False, indent=2)

def get_fund_codes(funds):
    return [f['code'] if isinstance(f, dict) else f for f in funds]

def find_fund_index(funds, code):
    for i, f in enumerate(funds):
        fund_code = f['code'] if isinstance(f, dict) else f
        if fund_code == code:
            return i
    return -1

@app.route('/')
def index():
    index_path = os.path.join(BASE_DIR, 'index.html')
    return send_file(index_path)

@app.route('/index.html')
def index_html():
    index_path = os.path.join(BASE_DIR, 'index.html')
    return send_file(index_path)

@app.route('/style.css')
def style_css():
    css_path = os.path.join(BASE_DIR, 'style.css')
    response = send_file(css_path, mimetype='text/css')
    # 支持带查询参数的请求（如缓存控制）
    if request.args.get('v') or request.args.get('t'):
        response.headers['Cache-Control'] = 'public, max-age=31536000'
    return response

@app.route('/app.js')
def app_js():
    js_path = os.path.join(BASE_DIR, 'app.js')
    response = send_file(js_path, mimetype='application/javascript')
    # 支持带查询参数的请求（如缓存控制）
    if request.args.get('v') or request.args.get('t'):
        response.headers['Cache-Control'] = 'public, max-age=31536000'
    return response

@app.route('/api/funds', methods=['GET'])
def get_funds():
    return jsonify(load_funds())

@app.route('/api/funds', methods=['POST'])
def add_fund():
    data = request.json
    code = data.get('code', '').strip()
    
    if not code or len(code) != 6 or not code.isdigit():
        return jsonify({'error': '请输入正确的6位基金代码'}), 400
    
    funds = load_funds()
    codes = get_fund_codes(funds)
    
    if code in codes:
        return jsonify({'error': '该基金已存在'}), 400
    
    funds.append({'code': code, 'shares': None})
    save_funds(funds)
    
    return jsonify({'success': True, 'code': code})

@app.route('/api/funds/<code>', methods=['DELETE'])
def remove_fund(code):
    funds = load_funds()
    index = find_fund_index(funds, code)
    
    if index >= 0:
        funds.pop(index)
        save_funds(funds)
        return jsonify({'success': True})
    
    return jsonify({'error': '基金不存在'}), 404

@app.route('/api/funds/<code>/shares', methods=['PUT'])
def update_fund_shares(code):
    data = request.json
    shares = data.get('shares')
    
    if shares is not None:
        try:
            shares = float(shares)
            if shares < 0:
                return jsonify({'error': '份额不能为负数'}), 400
        except (ValueError, TypeError):
            return jsonify({'error': '无效的份额数值'}), 400
    
    funds = load_funds()
    index = find_fund_index(funds, code)
    
    if index >= 0:
        funds[index]['shares'] = shares
        save_funds(funds)
        return jsonify({'success': True, 'code': code, 'shares': shares})
    
    return jsonify({'error': '基金不存在'}), 404

@app.route('/api/funds/order', methods=['PUT'])
def update_funds_order():
    data = request.json
    codes = data.get('funds', [])
    
    if not isinstance(codes, list):
        return jsonify({'error': '无效的数据格式'}), 400
    
    valid_codes = [c for c in codes if isinstance(c, str) and len(c) == 6 and c.isdigit()]
    
    if len(valid_codes) != len(codes):
        return jsonify({'error': '包含无效的基金代码'}), 400
    
    current_funds = load_funds()
    shares_map = {}
    for f in current_funds:
        if isinstance(f, dict):
            shares_map[f['code']] = f.get('shares')
    
    new_funds = []
    for code in valid_codes:
        new_funds.append({
            'code': code,
            'shares': shares_map.get(code)
        })
    
    save_funds(new_funds)
    return jsonify({'success': True})

@app.route('/api/export', methods=['POST'])
def export_funds():
    data = request.json
    funds_data = data.get('funds', [])
    
    if not funds_data:
        return jsonify({'error': '没有数据可导出'}), 400
    
    wb = Workbook()
    ws = wb.active
    ws.title = '基金列表'
    
    header_font = Font(bold=True, color='FFFFFF', size=12)
    header_fill = PatternFill(start_color='667EEA', end_color='667EEA', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    headers = ['基金名称', '基金代码', '净值日期', '单位净值', '基金份额', '基金市值', '最近分红']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    data_alignment = Alignment(horizontal='center', vertical='center')
    number_alignment = Alignment(horizontal='right', vertical='center')
    
    for row, fund in enumerate(funds_data, 2):
        ws.cell(row=row, column=1, value=fund.get('name', '-')).alignment = Alignment(horizontal='left', vertical='center')
        ws.cell(row=row, column=2, value=fund.get('code', '-')).alignment = data_alignment
        ws.cell(row=row, column=3, value=fund.get('netValueDate', '-')).alignment = data_alignment
        
        net_value = fund.get('netValue')
        if net_value:
            try:
                ws.cell(row=row, column=4, value=float(net_value)).alignment = number_alignment
                ws.cell(row=row, column=4).number_format = '#,##0.0000'
            except:
                ws.cell(row=row, column=4, value=net_value).alignment = data_alignment
        else:
            ws.cell(row=row, column=4, value='-').alignment = data_alignment
        
        shares = fund.get('shares')
        if shares:
            try:
                ws.cell(row=row, column=5, value=float(shares)).alignment = number_alignment
                ws.cell(row=row, column=5).number_format = '#,##0.00'
            except:
                ws.cell(row=row, column=5, value=shares).alignment = data_alignment
        else:
            ws.cell(row=row, column=5, value='-').alignment = data_alignment
        
        market_value = fund.get('marketValue')
        if market_value:
            try:
                ws.cell(row=row, column=6, value=float(market_value)).alignment = number_alignment
                ws.cell(row=row, column=6).number_format = '#,##0.00'
            except:
                ws.cell(row=row, column=6, value=market_value).alignment = data_alignment
        else:
            ws.cell(row=row, column=6, value='-').alignment = data_alignment
        
        ws.cell(row=row, column=7, value=fund.get('dividendDate', '-')).alignment = data_alignment
        
        for col in range(1, 8):
            ws.cell(row=row, column=col).border = thin_border
    
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 12
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'基金列表_{timestamp}.xlsx'
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

@app.route('/api/dividend/<code>', methods=['GET'])
def get_dividend(code):
    if not code or len(code) != 6 or not code.isdigit():
        return jsonify({'error': '无效的基金代码'}), 400
    
    dividend_date = fetch_dividend_date(code)
    if dividend_date:
        return jsonify({'code': code, 'dividendDate': dividend_date})
    else:
        return jsonify({'code': code, 'dividendDate': None})

@app.route('/api/fundname/<code>', methods=['GET'])
def get_fund_name(code):
    if not code or len(code) != 6 or not code.isdigit():
        return jsonify({'error': '无效的基金代码'}), 400
    
    name = fetch_fund_name(code)
    if name:
        return jsonify({'code': code, 'name': name})
    else:
        return jsonify({'code': code, 'name': None})

@app.route('/api/fundinfo/<code>', methods=['GET'])
def get_fund_info(code):
    if not code or len(code) != 6 or not code.isdigit():
        return jsonify({'error': '无效的基金代码'}), 400
    
    result = {
        'code': code,
        'name': None,
        'netValue': None,
        'netValueDate': None,
        'dayGrowth': None,
        'dividendDate': None
    }
    
    name = fetch_fund_name(code)
    if name:
        result['name'] = name
    
    fund_data = fetch_fund_realtime_data(code)
    if fund_data:
        if fund_data.get('name'):
            result['name'] = fund_data['name']
        result['netValue'] = fund_data.get('netValue')
        result['netValueDate'] = fund_data.get('netValueDate')
        result['dayGrowth'] = fund_data.get('dayGrowth')
    
    if not result['name']:
        name_from_page = fetch_fund_name_from_page(code)
        if name_from_page:
            result['name'] = name_from_page
    
    if not result['netValue']:
        history_data = fetch_fund_history_data(code)
        if history_data:
            result['netValue'] = history_data.get('netValue')
            result['netValueDate'] = history_data.get('netValueDate')
            result['dayGrowth'] = history_data.get('dayGrowth')
            if not result['name'] and history_data.get('name'):
                result['name'] = history_data['name']
    
    dividend = fetch_dividend_date(code)
    result['dividendDate'] = dividend
    
    return jsonify(result)

def fetch_fund_realtime_data(code):
    try:
        req = urllib.request.Request(
            f'https://fundgz.1234567.com.cn/js/{code}.js',
            headers={
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
                'Accept': '*/*',
                'Accept-Language': 'zh-CN,zh;q=0.9,en;q=0.8',
                'Referer': 'https://fund.eastmoney.com/'
            }
        )
        
        with urllib.request.urlopen(req, timeout=5) as response:
            content = response.read().decode('utf-8', errors='ignore')
            
            match = re.search(r'jsonpgz\s*\(\s*(\{.*?\})\s*\)', content)
            if match:
                try:
                    data = json.loads(match.group(1))
                    if data and data.get('fundcode'):
                        return {
                            'name': data.get('name'),
                            'netValue': data.get('dwjz'),
                            'netValueDate': data.get('jzrq'),
                            'dayGrowth': data.get('gszzl')
                        }
                except json.JSONDecodeError:
                    pass
    except Exception:
        pass
    
    return None

def fetch_fund_name_from_page(code):
    try:
        req = urllib.request.Request(
            f'https://fundf10.eastmoney.com/jjjz_{code}.html',
            headers={
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
                'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
                'Accept-Language': 'zh-CN,zh;q=0.9,en;q=0.8',
                'Referer': 'https://fund.eastmoney.com/'
            }
        )
        
        with urllib.request.urlopen(req, timeout=5) as response:
            html = response.read().decode('utf-8', errors='ignore')
            
            patterns = [
                r'<title>\s*([^<]+?)\s*\(',
                r'class="funCur"[^>]*>([^<]+)</a>',
                r'fundName\s*[=:]\s*["\']([^"\']+)["\']',
            ]
            
            for pattern in patterns:
                match = re.search(pattern, html)
                if match:
                    name = match.group(1).strip()
                    if name and len(name) > 2:
                        return name
    except Exception:
        pass
    
    return None

def fetch_fund_history_data(code):
    try:
        from datetime import datetime, timedelta
        
        end_date = datetime.now()
        start_date = end_date - timedelta(days=7)
        
        sdate = start_date.strftime('%Y-%m-%d')
        edate = end_date.strftime('%Y-%m-%d')
        
        req = urllib.request.Request(
            f'https://fund.eastmoney.com/f10/F10DataApi.aspx?type=lsjz&code={code}&page=1&sdate={sdate}&edate={edate}&per=10',
            headers={
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
                'Accept': '*/*',
                'Accept-Language': 'zh-CN,zh;q=0.9,en;q=0.8',
                'Referer': 'https://fund.eastmoney.com/'
            }
        )
        
        with urllib.request.urlopen(req, timeout=5) as response:
            content = response.read().decode('utf-8', errors='ignore')
            
            values = re.findall(r'<td>(\d{4}-\d{2}-\d{2})</td>\s*<td[^>]*>([\d.]+)</td>', content)
            
            if values:
                latest_date, latest_value = values[0]
                if len(values) > 1:
                    prev_date, prev_value = values[1]
                    try:
                        growth = (float(latest_value) - float(prev_value)) / float(prev_value) * 100
                        day_growth = f"{growth:.2f}"
                    except:
                        day_growth = None
                else:
                    day_growth = None
                
                return {
                    'netValue': latest_value,
                    'netValueDate': latest_date,
                    'dayGrowth': day_growth
                }
    except Exception:
        pass
    
    return None

FUND_NAME_CACHE = {}

def fetch_fund_name(code):
    global FUND_NAME_CACHE
    
    if code in FUND_NAME_CACHE:
        return FUND_NAME_CACHE[code]
    
    try:
        req = urllib.request.Request(
            f'https://fundgz.1234567.com.cn/js/{code}.js',
            headers={
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
                'Accept': '*/*',
                'Accept-Language': 'zh-CN,zh;q=0.9,en;q=0.8',
                'Referer': 'https://fund.eastmoney.com/'
            }
        )
        
        with urllib.request.urlopen(req, timeout=5) as response:
            content = response.read().decode('utf-8', errors='ignore')
            
            match = re.search(r'jsonpgz\s*\(\s*(\{.*?\})\s*\)', content)
            if match:
                try:
                    data = json.loads(match.group(1))
                    if data and data.get('name'):
                        FUND_NAME_CACHE[code] = data.get('name')
                        return data.get('name')
                except json.JSONDecodeError:
                    pass
    except Exception:
        pass
    
    return None

def fetch_dividend_date(code):
    urls_to_try = [
        f'https://fundf10.eastmoney.com/fhsp_{code}.html',
        f'https://fundf10.eastmoney.com/jjjz_{code}.html',
    ]
    
    for url in urls_to_try:
        try:
            req = urllib.request.Request(
                url,
                headers={
                    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
                    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
                    'Accept-Language': 'zh-CN,zh;q=0.9,en;q=0.8',
                    'Referer': 'https://fund.eastmoney.com/'
                }
            )
            
            with urllib.request.urlopen(req, timeout=5) as response:
                html = response.read().decode('utf-8', errors='ignore')
                
                patterns = [
                    r'(\d{4}-\d{2}-\d{2})\s*</td>\s*<td[^>]*>\s*\d{4}-\d{2}-\d{2}\s*</td>\s*<td[^>]*>每份派现金[\d.]+元',
                    r'(\d{4}-\d{2}-\d{2})\s*</td>\s*<td[^>]*>\s*\d{4}-\d{2}-\d{2}\s*</td>\s*<td[^>]*>每份派现金',
                    r'权益登记日[^\d]*(\d{4}-\d{2}-\d{2})',
                    r'除息日[^\d]*(\d{4}-\d{2}-\d{2})',
                ]
                
                for pattern in patterns:
                    matches = re.findall(pattern, html)
                    if matches:
                        valid_dates = []
                        for match in matches:
                            try:
                                date = datetime.strptime(match, '%Y-%m-%d')
                                if date.year >= 2015:
                                    valid_dates.append(match)
                            except ValueError:
                                continue
                        
                        if valid_dates:
                            valid_dates.sort(reverse=True)
                            return valid_dates[0]
                
                dividend_pattern = r'(\d{4})-(\d{2})-(\d{2})[^\d]*(?:每份派现金|分红|派现)'
                matches = re.findall(dividend_pattern, html)
                if matches:
                    valid_dates = []
                    for match in matches:
                        try:
                            date_str = f'{match[0]}-{match[1]}-{match[2]}'
                            date = datetime.strptime(date_str, '%Y-%m-%d')
                            if date.year >= 2015:
                                valid_dates.append(date_str)
                        except ValueError:
                            continue
                    
                    if valid_dates:
                        valid_dates.sort(reverse=True)
                        return valid_dates[0]
                        
        except Exception:
            continue
    
    return None

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)
